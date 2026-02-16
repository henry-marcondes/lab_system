from django.utils import timezone
from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render
from .models import Equipamento,  UsoEquipamento
from datetime import timedelta
from django.db.models import Q
# Excel
from openpyxl import Workbook
from django.http import HttpResponse

# PDF
from django.template.loader import get_template
from io import BytesIO
import csv



@login_required
def lista_equipamentos(request):
    equipamentos = Equipamento.objects.all()

    equipamentos_info = []

    for eq in equipamentos:
        uso_ativo = UsoEquipamento.objects.filter(
            equipamento=eq,
            data_devolucao__isnull=True
        ).first()

        pode_devolver = uso_ativo and uso_ativo.usuario == request.user

        equipamentos_info.append({
            'equipamento': eq,
            'uso_ativo': uso_ativo,
            'pode_devolver': pode_devolver
        })

    return render(request, 'equipamentos/lista.html', {
        'equipamentos_info': equipamentos_info
    })


@login_required
def usar_equipamento(request, equipamento_id):
    equipamento = get_object_or_404(Equipamento, id=equipamento_id)

    em_uso = UsoEquipamento.objects.filter(
        equipamento=equipamento,
        data_devolucao__isnull=True
    ).exists()

    if not em_uso:
        UsoEquipamento.objects.create(
            equipamento=equipamento,
            usuario=request.user
        )

        equipamento.status = Equipamento.STATUS_EM_USO   #  ESSENCIAL
        equipamento.save()                               #  ESSENCIAL

    return redirect('lista_equipamentos')


@login_required
def devolver_equipamento(request, equipamento_id):
    equipamento = get_object_or_404(Equipamento, id=equipamento_id)

    uso = UsoEquipamento.objects.filter(
        equipamento=equipamento,
        data_devolucao__isnull=True,
        usuario=request.user
    ).first()

    if uso and equipamento.status == Equipamento.STATUS_EM_USO:
        uso.data_devolucao = timezone.now()
        uso.save()

        equipamento.status = Equipamento.STATUS_DISPONIVEL
        equipamento.save()

    return redirect('lista_equipamentos')


@login_required
def equipamentos_em_uso(request):
    usos_ativos = UsoEquipamento.objects.filter(
        data_devolucao__isnull=True
    ).select_related('equipamento', 'usuario')

    agora = timezone.now()

    for uso in usos_ativos:
        diferenca = agora - uso.data_retirada
        horas, resto = divmod(diferenca.total_seconds(), 3600)
        minutos, _ = divmod(resto, 60)
        uso.tempo_em_uso = f"{int(horas)}h {int(minutos)}min"

    return render(request, 'equipamentos/em_uso.html', {
        'usos_ativos': usos_ativos
    })


@login_required
def historico_equipamento(request, equipamento_id):
    equipamento = get_object_or_404(Equipamento, id=equipamento_id)

    historico = UsoEquipamento.objects.filter(
        equipamento=equipamento
    ).select_related('usuario').order_by('-data_retirada')

    return render(request, 'equipamentos/historico_equipamento.html', {
        'equipamento': equipamento,
        'historico': historico
    })


@login_required
def meus_equipamentos(request):
    usos = UsoEquipamento.objects.filter(
        usuario=request.user,
        data_devolucao__isnull=True
    ).select_related('equipamento')

    return render(request, 'equipamentos/meus_equipamentos.html', {
        'usos': usos
    })


@login_required
def dashboard(request):
    total = Equipamento.objects.count()

    em_uso = Equipamento.objects.filter(status="em_uso").count()
    manutencao = Equipamento.objects.filter(status="manutencao").count()

    disponiveis = total - em_uso - manutencao

    meus_em_uso = UsoEquipamento.objects.filter(
        usuario=request.user,
        data_devolucao__isnull=True
    ).count()

    context = {
        "total": total,
        "disponiveis": disponiveis,
        "em_uso": em_uso,
        "manutencao": manutencao,
        "meus_em_uso": meus_em_uso,
    }

    return render(request, "equipamentos/dashboard.html", context)


@login_required
def manutencao(request, equipamento_id):
    equipamento = get_object_or_404(Equipamento, id=equipamento_id)

    print("status atual do equipamento:", equipamento.status)

    if equipamento.status != "disponivel":
        print("BLOQUEADO PELO IF")
        messages.error(request, "Equipamento não pode ir para manutenção.")
        return redirect("lista_equipamentos")

    equipamento.status = "manutencao"
    equipamento.save()
    print("STATUS NOVO:", equipamento.status)

    messages.success(request, "Equipamento enviado para manutenção.")
    return redirect("lista_equipamentos")


@login_required
def finalizar_manutencao(request, equipamento_id):
    equipamento = get_object_or_404(Equipamento, id=equipamento_id)

    if equipamento.status != "manutencao":
        messages.error(request, "Equipamento não está em manutenção.")
        return redirect("lista_equipamentos")

    equipamento.status = "disponivel"
    equipamento.save()

    messages.success(request, "Manutenção finalizada com sucesso.")
    return redirect("lista_equipamentos")


def relatorio_manutencao(request):
    equipamentos = Equipamento.objects.filter(status='manutencao')

    return render(
        request,
        "equipamentos/relatorio_manutencao.html",
        {"equipamentos": equipamentos}
    )


def relatorio_manutencao_csv(request):
    
    equipamentos = Equipamento.objects.filter(status='manutencao')

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        'attachment; filename="relatorio_manutencao.csv"'
    )

    writer = csv.writer(response)
    writer.writerow(["ID", "Nome", "Tipo", "Localização", "Status"])

    for eq in equipamentos:
        writer.writerow([
            eq.id,
            eq.nome,
            eq.tipo,
            eq.localizacao,
            eq.get_status_display()
        ])

    return response

def relatorio_manutencao_excel(request):
    
    equipamentos = Equipamento.objects.filter(status='manutencao')

    wb = Workbook()
    ws = wb.active
    ws.title = "Manutenção"

    ws.append(["ID", "Nome", "Tipo", "Localização", "Status"])

    for eq in equipamentos:
        ws.append([
            eq.id,
            eq.nome,
            eq.tipo,
            eq.localizacao,
            eq.get_status_display()
        ])

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )
    response["Content-Disposition"] = (
        'attachment; filename="relatorio_manutencao.xlsx"'
    )

    wb.save(response)
    return response

def relatorio_em_uso(request):
    usos = UsoEquipamento.objects.filter(
       data_devolucao__isnull=True
    ).select_related('equipamento', 'usuario')
    return render(request, "equipamentos/relatorio_em_uso.html", {"usos": usos})

def relatorio_em_uso_csv(request):
    usos = UsoEquipamento.objects.filter(data_devolucao__isnull=True).select_related('equipamento', 'usuario')

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        'attachment; filename="relatorio_em_uso.csv"'
    )

    writer = csv.writer(response)
    writer.writerow([
        "Equipamento", "Tipo","Usuário", "Data de Retirada"
    ])

    for uso in usos:
        writer.writerow([
            uso.equipamento.nome,
            uso.equipamento.tipo,
            uso.usuario.username,
            uso.data_retirada.strftime("%Y-%m-%d %H:%M:%S")
        ])

    return response

def relatorio_em_uso_excel(request):
    usos = UsoEquipamento.objects.filter(data_devolucao__isnull=True).select_related('equipamento', 'usuario')

    wb = Workbook()
    ws = wb.active
    ws.title = "Em Uso"

    ws.append([
        "Equipamento", "Tipo","Usuário", "Data de Retirada"
    ])

    for uso in usos:
        ws.append([
            uso.equipamento.nome,
            uso.equipamento.tipo,
            uso.usuario.username,
            uso.data_retirada.strftime("%Y-%m-%d %H:%M:%S")
        ])

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )
    response["Content-Disposition"] = (
        'attachment; filename="relatorio_em_uso.xlsx"'
    )

    wb.save(response)
    return response

def relatorio_por_usuario(request):
    usos = UsoEquipamento.objects.filter(
        data_devolucao__isnull=True
    ).select_related("equipamento", "usuario")

    return render(
        request,
        "equipamentos/relatorio_por_usuario.html",
        {"usos": usos}
    )

def relatorios_home(request):
    return render(request, "relatorios/home.html")

def relatorio_por_usuario_csv(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        'attachment; filename="relatorio_por_usuario.csv"'
    )

    writer = csv.writer(response)
    writer.writerow([
        "Usuário",
        "Equipamento",
        "Data de Retirada"
    ])

    usos = UsoEquipamento.objects.select_related(
        "usuario",
        "equipamento"
    )

    for uso in usos:
        writer.writerow([
            uso.usuario.username,
            uso.equipamento.nome,
            uso.data_retirada.strftime("%d/%m/%Y %H:%M")
        ])

    return response

def relatorio_por_usuario_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório por Usuário"

    # Cabeçalhos
    ws.append([
        "Usuário",
        "Equipamento",
        "Data de Retirada",
        "Data de Devolução"
    ])

    usos = UsoEquipamento.objects.select_related(
        "usuario",
        "equipamento"
    )

    for uso in usos:
        ws.append([
            uso.usuario.username,
            uso.equipamento.nome,
            uso.data_retirada.strftime("%d/%m/%Y %H:%M"),
            uso.data_devolucao.strftime("%d/%m/%Y %H:%M")
            if uso.data_devolucao else "Em uso"
        ])

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    )
    response[
        "Content-Disposition"
    ] = 'attachment; filename="relatorio_por_usuario.xlsx"'

    wb.save(response)
    return response





